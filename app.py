# -*- coding: utf-8 -*-
"""
Teklif → AX Dönüştürme API Servisi
Lovable uygulamasından gelen teklif Excel dosyasını okur,
orijinal formatta AX Excel dosyası üretir ve geri döner.
"""

import io
import os
import re
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)  # Lovable'dan gelen isteklere izin ver

# ─── Stiller ─────────────────────────────────────────────
RED = "FF0000"
BLUE = "0000FF"
BLACK = "000000"

thin = Side(style="thin", color=BLACK)
medium_blue = Side(style="medium", color=BLUE)


def cell_border(top=False, bottom=False, left=False, right=False):
    """İç ince kenarlık + dış kalın mavi kenarlık kombinasyonu"""
    return Border(
        top=medium_blue if top else thin,
        bottom=medium_blue if bottom else thin,
        left=medium_blue if left else thin,
        right=medium_blue if right else thin,
    )


# ─── Teklif Okuma ────────────────────────────────────────
def safe_filename(s):
    """Dosya adı için geçersiz karakterleri temizle."""
    return re.sub(r'[/\\:*?"<>|]', "-", s).strip()


def parse_teklif(file_bytes):
    """Teklif Excel'ini oku, üst bilgileri ve kalemleri ayıkla"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    info = {"magaza_adi": "", "magaza_kodu": "", "srv_no": ""}
    kalemler = []

    for row in ws.iter_rows(values_only=False):
        cells = [c.value for c in row]
        # Üst bilgileri bul: etiket hücresinin sağındaki dolu hücre değerdir
        for i, v in enumerate(cells):
            if v is None:
                continue
            text = str(v).strip()
            low = text.lower().replace(":", "").replace("ı", "i")
            if low == "magaza adi" or low == "mağaza adi" or "mağaza adı" in text.lower():
                info["magaza_adi"] = _next_value(cells, i)
            elif "mağaza kodu" in text.lower() or "magaza kodu" in low:
                info["magaza_kodu"] = _next_value(cells, i)
            elif low.startswith("srv no"):
                info["srv_no"] = _next_value(cells, i)

        # Kalem satırı mı? ET kodu içeren hücre ara
        et_idx = None
        for i, v in enumerate(cells):
            if v is not None and re.match(r"^ET\d+", str(v).strip()):
                et_idx = i
                break
        if et_idx is None:
            continue

        try:
            et_kod = str(cells[et_idx]).strip()
            aciklama = str(cells[et_idx + 1]).strip() if cells[et_idx + 1] else ""
            birim = str(cells[et_idx + 2]).strip() if cells[et_idx + 2] else ""
            miktar = _num(cells[et_idx + 3])
            birim_fiyat = _num(cells[et_idx + 4])
            tutar = _num(cells[et_idx + 5])
        except (IndexError, TypeError):
            continue

        # Sadece tutari 0'dan büyük kalemler AX'e girer
        if tutar and tutar > 0:
            kalemler.append({
                "aciklama": aciklama,
                "birim": birim,
                "miktar": miktar or 0,
                "birim_fiyat": birim_fiyat or 0,
                "et_kod": et_kod,
            })

    return info, kalemler


def _next_value(cells, start_idx):
    """Etiketten sonraki ilk dolu hücreyi döndür"""
    for j in range(start_idx + 1, len(cells)):
        if cells[j] is not None and str(cells[j]).strip() != "":
            return str(cells[j]).strip()
    return ""


def _num(v):
    """Hücre değerini sayıya çevir"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(".", "").replace(",", "."))
    except ValueError:
        return None


# ─── AX Dosyası Üretme ───────────────────────────────────
def build_ax(info, kalemler):
    """Orijinal formatta AX Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OZET"

    # Kılavuz çizgilerini kapat — orijinal AX görünümü
    ws.sheet_view.showGridLines = False
    # Sayfa Sonu Önizleme modu — orijinal AX'teki gri dış alan görünümü
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100
    ws.sheet_view.zoomScaleSheetLayoutView = 100

    # Sütun genişlikleri
    widths = {"A": 8, "B": 70, "C": 8, "D": 8, "E": 14, "F": 10, "G": 8, "H": 8, "I": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    n = len(kalemler)
    last_row = n + 1  # başlık + veri satırları

    # ── Başlık satırı ──
    headers = ["Poz no", "Açıklama", "Birim", "Miktar", "BF", "Para birimi", "", "", "Madde Kodu"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h if h else None)
        c.font = Font(name="Calibri", size=11, bold=True, color=RED)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = cell_border(
            top=True,
            bottom=(n == 0),
            left=(ci == 1),
            right=(ci == 9),
        )

    # ── Veri satırları ──
    for r_i, k in enumerate(kalemler):
        r = r_i + 2
        is_last = (r == last_row)

        vals = [
            r_i + 1,           # Poz no
            k["aciklama"],     # Açıklama
            "ADET",            # Birim — her zaman ADET
            k["miktar"],       # Miktar
            k["birim_fiyat"],  # BF
            "TRY",             # Para birimi
            None, None,        # boş sütunlar
            k["et_kod"],       # Madde Kodu
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name="Calibri", size=11)
            c.border = cell_border(
                bottom=is_last,
                left=(ci == 1),
                right=(ci == 9),
            )
            # Hizalama ve format
            if ci == 1:      # Poz no
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci == 3 or ci == 6:  # Birim, Para birimi
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 4:    # Miktar
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0.00"
            elif ci == 5:    # BF
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '"₺" #,##0.00'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # Yazdırma alanı — tablo bölgesi (grinin dışında kalan beyaz alan)
    ws.print_area = f"A1:I{last_row}"

    # Dosyayı belleğe kaydet
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════
# TEKLİF DOSYASI ÜRETME — Şablon tabanlı (logo şablondan korunur)
# ═════════════════════════════════════════════════════════
from datetime import datetime

SABLON_PATH = os.path.join(os.path.dirname(__file__), "sablon.xlsx")
LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")

# Şablondaki sabit satır yapısı
SEC_LAYOUT = {
    "A": {"baslik": 10, "veri": (11, 16), "toplam": 17},
    "B": {"baslik": 18, "veri": (19, 28), "toplam": 29},
    "C": {"baslik": 30, "veri": (31, 32), "toplam": 33},
    "D": {"baslik": 34, "veri": (35, 39), "toplam": 40},
    "E": {"baslik": 41, "veri": (42, 42), "toplam": 43},
}


def build_teklif(data):
    """Orijinal şablonu açar, sadece verileri doldurur. Logo şablonda gömülü kalır."""
    wb = load_workbook(SABLON_PATH)
    ws = wb.active

    # ── Logo: şablonun resmine güvenme (Python 3.14 openpyxl düşürüyor), koddan ekle ──
    try:
        ws._images = []  # şablondan gelen (bozulmuş olabilecek) resmi temizle
    except Exception:
        pass
    if os.path.exists(LOGO_PATH):
        try:
            from openpyxl.drawing.image import Image as XLImage
            logo_img = XLImage(LOGO_PATH)
            logo_img.width = 239
            logo_img.height = 70
            ws.add_image(logo_img, "B1")
        except Exception:
            pass

    # ── Üst bilgi ──
    ws["F1"] = data.get("magaza_adi", "")
    ws["F2"] = data.get("magaza_kodu", "")
    ws["F3"] = data.get("srv_no", "")
    ws["F4"] = data.get("tarih", datetime.now().strftime("%d.%m.%Y"))
    ws["F5"] = "TRY"
    ws["F6"] = data.get("hazirlayan", "YUSUF FIRAT YAY")

    # ── Şablondaki eski kalem verilerini temizle (ET kod/açıklama dahil) ──
    for sec, lay in SEC_LAYOUT.items():
        for r in range(lay["veri"][0], lay["veri"][1] + 1):
            for col in ("A", "B", "C", "D", "E", "F", "G", "I"):
                ws[f"{col}{r}"] = None

    # ── Kalemleri bölümlere göre grupla ──
    by_sec = {}
    for k in data.get("kalemler", []):
        by_sec.setdefault(k.get("bolum", "E"), []).append(k)

    # ── Kalemleri sırayla boş satırlara yaz ──
    for sec in ["A", "B", "C", "D", "E"]:
        items = by_sec.get(sec, [])
        lay = SEC_LAYOUT[sec]
        v_start, v_end = lay["veri"]
        no = 1
        r = v_start
        for k in items:
            if r > v_end:
                break
            ws[f"A{r}"] = no
            ws[f"B{r}"] = str(k.get("et_kodu", "")).strip()
            ws[f"C{r}"] = k.get("aciklama", "")
            ws[f"D{r}"] = k.get("birim", "ADET")
            ws[f"E{r}"] = k.get("miktar", 0)
            ws[f"F{r}"] = k.get("birim_fiyat", 0)
            ws[f"G{r}"] = f"=E{r}*F{r}"
            if k.get("not"):
                ws[f"I{r}"] = k.get("not")
            no += 1
            r += 1

    # ── Bölüm ara toplamları ve genel toplamlar (formül) ──
    for sec, lay in SEC_LAYOUT.items():
        v_start, v_end = lay["veri"]
        ws[f"H{lay['toplam']}"] = f"=SUM(G{v_start}:G{v_end})"

    toplam_satirlari = [SEC_LAYOUT[s]["toplam"] for s in SEC_LAYOUT]
    toplam_formula = "+".join([f"H{r}" for r in toplam_satirlari])
    ws["H53"] = f"={toplam_formula}"
    ws["H54"] = None
    ws["H55"] = "=H53*0.2"
    ws["H56"] = "=H53+H55"

    # ── Notlar ──
    notlar = data.get("notlar", {})
    if notlar.get("uretim_sevk_suresi") not in (None, ""):
        ws["E45"] = notlar.get("uretim_sevk_suresi")
    if notlar.get("tadilat_suresi") not in (None, ""):
        ws["E46"] = notlar.get("tadilat_suresi")
    if notlar.get("not3"):
        ws["C47"] = notlar.get("not3")
    if notlar.get("not4"):
        ws["C48"] = notlar.get("not4")

    # ── Ek not (yoksa şablondaki eskiyi temizle) ──
    ws["C52"] = data.get("ek_not", "") or None

    # Excel açar açmaz formülleri hesaplasın
    wb.calculation.fullCalcOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/teklif", methods=["POST"])
def create_teklif():
    """JSON verisinden teklif Excel'i üret"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "JSON verisi bulunamadı."}), 400

    if not data.get("magaza_adi") or not data.get("magaza_kodu"):
        return jsonify({"error": "Mağaza adı ve kodu zorunludur."}), 400

    try:
        buf = build_teklif(data)
    except Exception as e:
        return jsonify({"error": f"Teklif oluşturulamadı: {str(e)}"}), 500

    parts = []
    if data.get("magaza_kodu"):
        parts.append(data["magaza_kodu"])
    if data.get("magaza_adi"):
        parts.append(data["magaza_adi"])
    if data.get("srv_no"):
        parts.append(data["srv_no"])
    fname = safe_filename(" ".join(parts)) + " - TEKLIF.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/teklif-ve-ax", methods=["POST"])
def create_teklif_and_ax():
    """JSON verisinden hem teklif hem AX üret, ikisini zip olarak döndür"""
    import zipfile as zf

    data = request.get_json()
    if not data:
        return jsonify({"error": "JSON verisi bulunamadı."}), 400

    try:
        teklif_buf = build_teklif(data)
        # AX için kalemleri dönüştür (tutar > 0 olanlar)
        info = {
            "magaza_adi": data.get("magaza_adi", ""),
            "magaza_kodu": data.get("magaza_kodu", ""),
            "srv_no": data.get("srv_no", ""),
        }
        ax_kalemler = []
        for k in data.get("kalemler", []):
            tutar = (k.get("miktar", 0) or 0) * (k.get("birim_fiyat", 0) or 0)
            if tutar > 0:
                ax_kalemler.append({
                    "aciklama": k.get("aciklama", ""),
                    "birim": k.get("birim", "ADET"),
                    "miktar": k.get("miktar", 0),
                    "birim_fiyat": k.get("birim_fiyat", 0),
                    "et_kod": k.get("et_kodu", ""),
                })
        ax_buf = build_ax(info, ax_kalemler)
    except Exception as e:
        return jsonify({"error": f"Dosyalar oluşturulamadı: {str(e)}"}), 500

    # Dosya isimleri: TEKLİF = "kod-ad-srv", AX = "AX-kod-ad-srv"
    parcalar = [p for p in [data.get("magaza_kodu",""), data.get("magaza_adi",""), data.get("srv_no","")] if p]
    base = safe_filename("-".join(parcalar))
    teklif_name = f"{base}.xlsx"
    ax_name = f"AX-{base}.xlsx"

    zip_buf = io.BytesIO()
    with zf.ZipFile(zip_buf, "w", zf.ZIP_DEFLATED) as z:
        z.writestr(teklif_name, teklif_buf.read())
        z.writestr(ax_name, ax_buf.read())
    zip_buf.seek(0)

    return send_file(
        zip_buf,
        as_attachment=True,
        download_name=f"{base}.zip",
        mimetype="application/zip",
    )


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/ax", methods=["POST"])
def convert_ax():
    """Teklif Excel'i al, AX Excel'i döndür"""
    if "file" not in request.files:
        return jsonify({"error": "Dosya bulunamadı. 'file' alanında Excel gönderin."}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Sadece .xlsx veya .xlsm dosyaları kabul edilir."}), 400
    try:
        file_bytes = f.read()
        info, kalemler = parse_teklif(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Dosya okunamadı: {str(e)}"}), 422
    if not kalemler:
        return jsonify({"error": "Dosyada tutarı 0'dan büyük kalem bulunamadı."}), 422
    ax_buf = build_ax(info, kalemler)
    parts = ["AX"]
    if info["magaza_kodu"]:
        parts.append(info["magaza_kodu"])
    if info["magaza_adi"]:
        parts.append(info["magaza_adi"])
    if info["srv_no"]:
        parts.append(info["srv_no"])
    fname = safe_filename("-".join(parts)) + ".xlsx"
    return send_file(ax_buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/ax/preview", methods=["POST"])
def preview_ax():
    """İndirmeden önce önizleme verisi döndür (JSON)"""
    if "file" not in request.files:
        return jsonify({"error": "Dosya bulunamadı."}), 400
    f = request.files["file"]
    try:
        file_bytes = f.read()
        info, kalemler = parse_teklif(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Dosya okunamadı: {str(e)}"}), 422
    return jsonify({"info": info, "kalemler": kalemler, "kalem_sayisi": len(kalemler)})


@app.route("/diag", methods=["GET"])
def diag():
    """Teşhis: Python sürümü, logo durumu, dosya yolları"""
    import sys
    info = {
        "python": sys.version,
        "logo_path": LOGO_PATH,
        "logo_exists": os.path.exists(LOGO_PATH),
        "logo_size": os.path.getsize(LOGO_PATH) if os.path.exists(LOGO_PATH) else 0,
        "sablon_path": SABLON_PATH,
        "sablon_exists": os.path.exists(SABLON_PATH),
        "cwd": os.getcwd(),
        "dir_listing": os.listdir(os.path.dirname(os.path.abspath(__file__))),
    }
    # Logo eklenebilir mi test et
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        twb = Workbook()
        tws = twb.active
        if os.path.exists(LOGO_PATH):
            img = XLImage(LOGO_PATH)
            tws.add_image(img, "A1")
            import io as _io
            b = _io.BytesIO()
            twb.save(b)
            b.seek(0)
            import zipfile as _zf
            with _zf.ZipFile(b) as _z:
                info["test_logo_media"] = [n for n in _z.namelist() if "media" in n]
        info["add_image_works"] = True
    except Exception as e:
        info["add_image_works"] = False
        info["add_image_error"] = str(e)
    return jsonify(info)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
